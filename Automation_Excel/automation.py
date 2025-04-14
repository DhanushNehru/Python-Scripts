from typing import ValuesView
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def processing(filename):
  wb = xl.load_workbook(filename)
  sheet = wb['Sheet1']
  cell = sheet.cell(1, 2)
  print(sheet.max_row)
  for i in range(2, sheet.max_row+1):
         print(sheet.cell(i, 3).value)
  new_cell = sheet.cell(1, 4)
  new_cell.value = "new cell"
  for i in range(2, 5):
      new_values = sheet.cell(i, 4)
      new_values.value = int(input(""))
  Values = Reference(
    sheet,
    min_row = 1,
    max_row = 5,
    min_col = 1,
    max_col = 5
  )
  chart = BarChart()
  chart.add_data(Values)
  sheet.add_chart(chart, 'a5')
  wb.save(filename)


filename = input("enter the name of the file: ")
processing(filename)
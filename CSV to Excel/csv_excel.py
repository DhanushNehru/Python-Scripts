import openpyxl
import os
from openpyxl.styles import Font, PatternFill

# Get the CSV and Excel file names from the user
csv_files = input("Enter the CSV files separated by commas (e.g., file1.csv, file2.csv): ").split(',')
sep = input("Separator of the CSV files (default is ','): ") or ','  # Default to comma separator if not provided
excel_name = input("Name of the output Excel file with extension: ")

# Load or create Excel workbook
if os.path.exists(excel_name):
    workbook = openpyxl.load_workbook(excel_name)
else:
    workbook = openpyxl.Workbook()

# Loop over multiple CSV files to write them into different sheets
for csv_name in csv_files:
    csv_name = csv_name.strip()  # Trim any whitespace
    sheet_name = os.path.splitext(os.path.basename(csv_name))[0]  # Sheet name based on the CSV filename

    # Create a new sheet for each CSV file
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)

    # Write CSV data to the Excel sheet
    try:
        with open(csv_name, "r", encoding="utf-8") as file:
            excel_row = 1
            header_detected = False  # Flag to check if header formatting should be applied

            for line in file:
                data = line.strip().split(sep)
                excel_column = 1

                # Apply header formatting for the first row (headers)
                if not header_detected:
                    for value in data:
                        cell = sheet.cell(row=excel_row, column=excel_column, value=value)
                        # Apply bold font and background color for the header row
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        excel_column += 1
                    header_detected = True  # Mark the first row as header
                else:
                    for value in data:
                        sheet.cell(row=excel_row, column=excel_column, value=value)
                        excel_column += 1

                excel_row += 1

    except FileNotFoundError:
        print(f"Error: The CSV file '{csv_name}' was not found.")
    except Exception as e:
        print(f"An error occurred while processing {csv_name}: {e}")

# Save the Excel file with all sheets
workbook.save(excel_name)

print(f"All CSV files have been processed and saved to {excel_name}.")


import openpyxl
from datetime import datetime
from tkinter import *
from tkinter import filedialog, messagebox
from openpyxl.utils import get_column_letter
from openpyxl import Workbook


class CalculateApp:

    def __init__(self, root):
        self.root = root
        self.file_entry = StringVar()
        head_label = Label(root, text="Let's calculate the payment and hours",
                           padx=15,
                           pady=15,
                           font="SegoeUI 14",
                           bg="palegreen1",
                           fg="red")
        head_label.grid(row=1,
                        column=1,
                        pady=10,
                        padx=5,
                        columnspan=3)

        file_label = Label(root, text="XL File Path:", font="SegoeUI 12", bg="PaleGreen1")
        file_label.grid(row=2, column=1, pady=10, padx=5)

        self.file_entry = Entry(root, width=30)
        self.file_entry.grid(row=2, column=2, pady=10, padx=5)

        browse_button = Button(root, text="Browse", font="SegoeUI 12", bg="PaleGreen1", command=self.browse)
        browse_button.grid(row=2, column=3, pady=10, padx=5)

        pay_label = Label(root, text="Pay per Hour:", font="SegoeUI 12", bg="PaleGreen1")
        pay_label.grid(row=3, column=1, pady=10, padx=5)

        self.pay_entry = Entry(root, width=30)
        self.pay_entry.grid(row=3, column=2, pady=10, padx=5)

        calculate_button = Button(root, text="Calculate", font="SegoeUI 12", bg="PaleGreen1",
                                  command=self.calculate_button_clicked)
        calculate_button.grid(row=4, column=1, pady=10, padx=5, columnspan=3)

    def browse(self):
        # Presenting the user with a pop-up for directory selection
        file_path = filedialog.askopenfilename(title="Select XL File")
        # Checking if the selected file has a valid Excel file extension
        if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
            self.file_entry.delete(0, END)  # Clearing any existing text
            self.file_entry.insert(0, file_path)
        else:
            self.file_entry.delete(0, END)
            self.file_entry.insert(0, "Invalid file format")

    def calculate_payment(self, file_path, pay_per_hour):
        try:
            # Implement your logic to calculate the payment based on the hours in the xl file
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            total_price = 0
            for row in range(2, sheet.max_row + 1):
                cellStart = sheet.cell(row, 1)
                cellFinish = sheet.cell(row, 2)
                start_time = datetime.combine(datetime.now().date(), cellStart.value)
                finish_time = datetime.combine(datetime.now().date(), cellFinish.value)
                total_time = finish_time - start_time
                total_time_cell = sheet.cell(row, 3)
                total_time_cell.value = total_time.total_seconds() / 3600
                price = pay_per_hour * (total_time.total_seconds() / 3600)
                price_cell = sheet.cell(row, 4)
                price_cell.value = price
                price_cell.number_format = '0.00'  # Set the number format to display as a double in shekels
                total_price += price

            total_price_cell = sheet.cell(sheet.max_row + 1, 4)  # Place the total price in the fifth cell
            total_price_cell.value = total_price
            total_price_cell.number_format = '0.00'  # Set the number format to display as a double in shekels
            # Save the updated Excel file
            modified_file_path = f"modified_{file_path}"
            wb.save(modified_file_path)

            messagebox.showinfo("Calculation Complete", f"Total Payment: {total_price:.2f} Shekels")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def calculate_button_clicked(self):
        file_path = self.file_entry.get()
        pay_per_hour = self.pay_entry.get()
        if not file_path:
            messagebox.showerror("Error", "Please select an Excel file.")
        elif not pay_per_hour:
            messagebox.showerror("Error", "Please enter the pay per hour.")
        else:
            try:
                pay_per_hour = float(pay_per_hour)
                self.calculate_payment(file_path, pay_per_hour)
            except ValueError:
                messagebox.showerror("Error", "Invalid pay per hour value. Please enter a number.")


def main():
    root = Tk()
    # Set the title, background color, and size of the tkinter window and disabling the resizing property
    root.geometry("520x280")
    root.resizable(False, False)
    root.title("Freelance help software")
    root.config(background="PaleGreen1")
    # Creating an instance of CalculateApp class
    app = CalculateApp(root)
    # Running the application
    root.mainloop()


if __name__ == '__main__':
    main()
